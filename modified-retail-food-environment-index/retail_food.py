
import os
import sys
import traceback

from openpyxl import load_workbook
from sqlalchemy import create_engine

sys.path.append('..')

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def db_exec(eng, this_sql):
    # print(f"sql: {sql}")
    if this_sql.strip().startswith('select'):
        return [dict(row) for row in eng.execute(this_sql).fetchall()]
    else:
        # print(f"sql: {this_sql}")
        return eng.execute(this_sql)
        foo = intput()

def fix(val):
    if val is None or val == '':
        return 'NULL'
    else:
        val = str(val).replace("'", "''").replace('%', '%%')
        return f"'{val}'"


def fix_int(val):
    if val is None or val == '':
        return 'NULL'
    else:
        return val


if __name__ == '__main__':

    db_exec(conn, "drop table if exists modified_retail_food_env")

    print("\ncreating table modified_retail_food_env...")

    sql = """
          create table modified_retail_food_env (
              ind_id varchar(63),
              ind_definition varchar(63),
              reportyear varchar(63),
              race_eth_code varchar(63),
              race_eth_name varchar(63),
              geotype varchar(63),
              geotypevalue varchar(63),
              geoname varchar(63),
              county_name varchar(63),
              county_fips varchar(63),
              region_name varchar(63),
              region_code varchar(63),
              strata_one_code varchar(63),
              strata_one_name varchar(63),
              strata_two_code int,
              strata_two_name varchar(63),
              numerator int,
              denominator int,
              estimate float,
              ll_95ci float,
              ul_95ci float,
              se float,
              rse float,
              ca_decile int,
              ca_rr float,
              version varchar(63))"""

    db_exec(conn, sql)

    wb = load_workbook('modified-retail-food-environment-index-data.xlsx')

    ws = wb.active

    num = 2

    print("\nadding data...")

    while True:
  
        A = f"A{num}"
        B = f"B{num}"
        C = f"C{num}"
        D = f"D{num}"
        E = f"E{num}"
        F = f"F{num}"
        G = f"G{num}"
        H = f"H{num}"
        I = f"I{num}"
        J = f"J{num}"
        K = f"K{num}"
        L = f"L{num}"
        M = f"M{num}"
        N = f"N{num}"
        O = f"O{num}"
        P = f"P{num}"
        Q = f"Q{num}"
        R = f"R{num}"
        S = f"S{num}"
        T = f"T{num}"
        U = f"U{num}"
        V = f"V{num}"
        W = f"W{num}"
        X = f"X{num}"
        Y = f"Y{num}"
        Z = f"Z{num}"

        if ws[A].value is None and ws[B].value is None:
            print(f"\nadded rows # {num - 2}")
            break

        sql = f"""insert into modified_retail_food_env
                  values
                  ({fix(ws[A].value)},
                   {fix(ws[B].value)},
                   {fix(ws[C].value)},
                   {fix(ws[D].value)},
                   {fix(ws[E].value)},
                   {fix(ws[F].value)},
                   {fix(ws[G].value)},
                   {fix(ws[H].value)},
                   {fix(ws[I].value)},
                   {fix(ws[J].value)},
                   {fix(ws[K].value)},
                   {fix(ws[L].value)},
                   {fix(ws[M].value)},
                   {fix(ws[N].value)},
                   {fix_int(ws[O].value)},
                   {fix(ws[P].value)},
                   {fix_int(ws[Q].value)},
                   {fix_int(ws[R].value)},
                   {fix_int(ws[S].value)},
                   {fix_int(ws[T].value)},
                   {fix_int(ws[U].value)},
                   {fix_int(ws[V].value)},
                   {fix_int(ws[W].value)},
                   {fix_int(ws[X].value)},
                   {fix_int(ws[Y].value)},
                   {fix(ws[Z].value)})"""

        # print(f"sql: {sql}")
        db_exec(conn, sql)

        num += 1
