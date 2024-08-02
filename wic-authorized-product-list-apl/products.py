
import os
import traceback

from openpyxl import load_workbook
from sqlalchemy import create_engine

sys.path.append('..')

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def db_exec(eng, this_sql):
    # print(f"sql: {sql}")
    if this_sql.strip().startswith('select'):
        return [dict(row) for row in eng.execute(this_sql).fetchall()]
    else:
        # print(f"sql: {this_sql}")
        return eng.execute(this_sql)
        foo = intput()

def fix(val):
    if val is None or val == '':
        return 'NULL'
    else:
        val = val.replace("'", "''").replace('%', '%%')
        return f"'{val}'"

def fix_int(val):
    if val is None or val == '':
        return 'NULL'
    else:
        return val


def is_valid_upc(s):
    return len(s) >= 9


if __name__ == '__main__':

    f = None

    for file in os.listdir():
        if file.endswith('xlsx'):
            f = file
            break

    if f is None:
        raise Exception(f"No xlsx file found in files: {os.listdir()}")
    else:
        print(f"found file: {f}")

    db_exec(conn, "drop table if exists wic_products")

    sql = """
          create table wic_products (
              upc varchar(15),
              category char(2),
              cat_desc varchar(127),
              sub_category char(3),
              sub_cat_desc varchar(127),
              prod_desc_and_size varchar(127),
              uom varchar(5))"""
    db_exec(conn, sql)

    wb = load_workbook('ca-wic-apl-07-05-2024.xlsx')

    ws = wb.active

    num = 2
    done = False

    while not done:
  
        A = f"A{num}"
        B = f"B{num}"
        C = f"C{num}"
        D = f"D{num}"
        E = f"E{num}"
        F = f"F{num}"
        G = f"G{num}"

        if not is_valid_upc(str(ws[A].value)):
            done = True
            continue

        upc = str(ws[A].value)
        category = str(ws[B].value).zfill(2)
        sub_category = str(ws[D].value).zfill(3)

        sql = f"""insert into wic_products
                  (upc, category, cat_desc, sub_category, sub_cat_desc, prod_desc_and_size, uom)
                  values
                  ({fix(upc)},
                   {fix(category)},
                   {fix(ws[C].value)},
                   {fix(sub_category)},
                   {fix(ws[E].value)},
                   {fix(ws[F].value)},
                   {fix(ws[G].value)})"""
        # print(f"sql: {sql}")
        db_exec(conn, sql)

        num += 1
