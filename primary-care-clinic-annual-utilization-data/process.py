
import argparse
import os
import re

import traceback

from openpyxl import load_workbook
from sqlalchemy import create_engine

sys.path.append('..')

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def arguments():
    parser = argparse.ArgumentParser()

    parser.add_argument('--list-columns', action='store_true')
    parser.add_argument('--facilities', action='store_true')

    return parser.parse_args()


def db_exec(engine, sql):
    # print(f"sql: {sql}")
    if sql.strip().startswith('select'):
        return [dict(r) for r in engine.execute(sql).fetchall()]
    else:
        return engine.execute(sql)


def fetch_max_pk(table):
    sql = f"select max(pk) as pk from {table}"
    result = db_exec(conn, sql)[0]['pk']
    if result is None:
        return 0
    else:
        return result


def fix_value(start):
    if start is None:
        return 'NULL'
    else:
        start = str(start).replace("'", "''")
        return f"'{start}'"


def find_facilities_files():

    pk = fetch_max_pk('pccu_files')

    for f in os.listdir('.'):

        if not re.search('-utilization-data', str(f)) or re.search('-pivot-', str(f)):
            continue

        if not str(f).endswith('.xlsx'):
            continue

        pk += 1
        sql = f"insert into pccu_files values ({pk}, '{f}', 'utilization')"
        db_exec(conn, sql)


if __name__ == '__main__':

    args = arguments()

    if args.facilities:

        find_facilities_files()

        fac_columns = [
            'fac_acquire_equip_over_500k',
            'fac_address_one',
            'fac_address_two',
            'fac_admin_name',
            'fac_city',
            'fac_name',
            'fac_no',
            'fac_operated_this_yr',
            'fac_oper_curryr',
            'fac_op_per_begin_dt',
            'fac_op_per_end_dt',
            'fac_par_corp_bus_addr',
            'fac_par_corp_city',
            'fac_par_corp_name',
            'fac_par_corp_state',
            'fac_par_corp_zip',
            'fac_phone',
            'fac_str_addr',
            'fac_zip',
            'fac_zipcode',
            'oshpd_id',
            'parent_address_one',
            'parent_address_two',
            'parent_city',
            'parent_name',
            'parent_state',
            'parent_zipcode',
            'pcc_health_serv_medical',
            'pcc_health_serv_dental',
            'pcc_health_serv_vision',
            'pcc_health_serv_mental_health',
            'pcc_health_serv_substance_abuse',
            'pcc_health_serv_domestic_violence',
            'pcc_health_serv_basic_lab',
            'pcc_health_serv_radiological_services',
            'pcc_health_serv_urgent_care',
            'pcc_health_serv_pharmacy',
            'pcc_health_serv_womens_health']
 
        bool_columns = [
            'fac_acquire_equip_over_500k',
            'fac_operated_this_yr',
            'fac_oper_curryr',
            'pcc_health_serv_medical',
            'pcc_health_serv_dental',
            'pcc_health_serv_vision',
            'pcc_health_serv_mental_health',
            'pcc_health_serv_substance_abuse',
            'pcc_health_serv_domestic_violence',
            'pcc_health_serv_basic_lab',
            'pcc_health_serv_radiological_services',
            'pcc_health_serv_urgent_care',
            'pcc_health_serv_pharmacy',
            'pcc_health_serv_womens_health']

        int_columns = ['pk', 'fac_no', 'oshpd_id']

        pk = 0

        sql = "select * from pccu_files where tag = 'utilization'"

        for frow in db_exec(conn, sql):

            fn = frow['name']

            wb = load_workbook(fn)

            fac_column_defs = dict()

            for sheetname in wb.sheetnames:

                if not sheetname.startswith('Section ') and not sheetname.startswith('Page '):
                    continue

                wb.active = wb[sheetname]

                row_num = 1
                col_num = 1

                done_reading_col_heads = False
                max_column_num = 1

                while not done_reading_col_heads:

                    value = wb.active.cell(row_num, col_num).value

                    if value is None or value == '':
                        done_reading_col_heads = True
                    else:
                        value = value.strip().lower()
                        if value in fac_columns:
                            fac_column_defs[value] = col_num
                        col_num += 1
                        max_column_num = col_num

                print(f"file: {fn}")
                print(f"sheetname: {sheetname}")
                # print(f"fac_column_defs: {fac_column_defs}")

                if len(fac_column_defs) > 0:

                    row_num = 5

                    done_reading_rows = False

                    while not done_reading_rows:

                        a1 = wb.active.cell(row_num, 1).value
                        if a1 in ['Page', 'Column', 'Line']:
                            row_num += 1
                            continue

                        next_row = dict()

                        keys = list(fac_column_defs.keys())

                        for key in keys:
                            value = wb.active.cell(row_num, fac_column_defs[key]).value

                            if value is not None and value != '':
                                next_row[key] = fix_value(value)

                                if key in bool_columns:
                                    next_row[key] = next_row[key].replace("'Yes'", "'Y'").replace("'No'", "'N'")

                                if key in int_columns:
                                    next_row[key] = next_row[key].replace("'","")

                        if 'oshpd_id' in next_row and 'fac_no' not in next_row:
                            next_row['fac_no'] = next_row.pop('oshpd_id')
 
                        if len(next_row) == 0:
                            done_reading_rows = True
                            continue

                        next_row['sheetname'] = f"'{str(sheetname)}'"
                        next_row['file_pk'] = str(frow['pk'])

                        pk += 1
                        next_row['pk'] = str(pk)

                        # print(f"next_row: {next_row}")

                        cols = ', '.join([r for r in list(next_row.keys())])

                        vals = ', '.join([r for r in list(next_row.values())])
 
                        sql = f"insert ignore into pccu_facilities ({cols}) values ({vals})"

                        # print(f"sql: {sql}")
                        db_exec(conn, sql)

                        row_num += 1

        # fix columns with empty values from the spreadsheet.
        for col in fac_columns:
            if col not in int_columns:
                sql = f"update pccu_facilities set {col} = NULL where {col} in ('', '0')"
                db_exec(conn, sql)


    if args.list_columns:

        fs = dict()

        for f in files():

            wb = load_workbook(f)

            fn = str(f)

            fs[fn] = dict()

            for sn in wb.sheetnames:

                if not sn.startswith('Section ') and not sn.startswith('Page '):
                    continue

                fs[fn][sn] = list()

                wb.active = wb[sn]

                col_num = 1
                done = False

                while not done:

                    value = wb.active.cell(1, col_num).value
                    if not value or value == '':
                        done = True
                    else:
                        # print(f"        A{col_num}: {value}")
                        fs[fn][sn].append(value.strip().lower())
                        col_num += 1

                # print(f"fs[{fn}] = {fs[fn]}")

        cols = dict()
    
        for fn in fs:
            for sn in fs[fn]:
                for label in fs[fn][sn]:
                    if label not in cols:
                        cols[label] = list()
                    cols[label].append(fn)

        print("")
        for col in sorted(cols):
            print(f"col: {col}")
            print(f"files:")
            for fn in cols[col]:
                print(f"    {fn}")
            print("")
 
        quit()


