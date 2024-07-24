
import argparse
import re
import traceback

from openpyxl import load_workbook
from sqlalchemy import create_engine

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def arguments():
    parser = argparse.ArgumentParser()

    parser.add_argument('--only-load-strings', action='store_true')
    parser.add_argument('--update-charge', action='store_true')
    parser.add_argument('--update-files-counts', action='store_true')
    parser.add_argument('--file', nargs='?', help="Name of single file to process.")
    parser.add_argument('--files', nargs='?', help="File containing list of files to process.")
    parser.add_argument('--year', nargs='?', help="Files only for this year will be processed.")
    parser.add_argument('--verbose', '-v', action='store_true')

    return parser.parse_args()


def db_exec(engine, sql):
    # print(f"sql: {sql}")
    if sql.strip().startswith('select'):
        return [dict(r) for r in engine.execute(sql).fetchall()]
    else:
        return engine.execute(sql)


def is_cpt_label(cell):
    if cell.value is None:
        return False
    if cell.data_type != 's':
        return False
    if not cell.value.endswith('CPT Code'):
        return False
    return True


def fix(val):
     if val is None:
         return "NULL"
     else:
         val = str(val).replace("'", "''")
         return f"'{val}'"


def common25_sheet(wbook):
    for found_name in wbook.sheetnames:
        if 'Common' in found_name or 'COMMON' in found_name or 'AB 1045' in found_name:
            return wbook[found_name]
    return None

def dprint(s):
    if args.verbose:
        print(s)


def load_strings():

    if not incremental:

        print("re-creating table; chargemasters_common25")
        db_exec(conn, "drop table if exists chargemasters_common25")
        sql = """
              create table chargemasters_common25 (
                  file_pk int,
                  procedure_desc varchar(1027),
                  cpt_code varchar(31),
                  charge_str varchar(31),
                  charge decimal(10,2))"""
        db_exec(conn, sql)
        db_exec(conn, "alter table chargemasters_common25 add index (charge_str)")
        db_exec(conn, "alter table chargemasters_common25 add index (charge)")
        db_exec(conn, "alter table chargemasters_common25 add index (file_pk)")

        sql = """
              select pk, full_name, file_type from chargemasters_files
              where file_type in ('Common25', 'CDM All') and file_ext = 'xlsx'"""
        files = db_exec(conn, sql)
    if args.file is not None:
        sql = f"select pk, full_name, file_type from chargemasters_files where full_name = {fix(args.file)} and file_ext = 'xlsx'"
        files = db_exec(conn, sql)
    if args.files is not None:
        fs = open(args.files, 'r')
        fnames = fs.readlines()
        fnames = [fix(r) for r in fnames]
        sql = f"select pk, full_name, file_type from chargemasters_files where full_name in ({','.join(fnames)}) and file_ext = 'xlsx'"
        files = db_exec(conn, sql)
    if args.year is not None:
        sql = f"select pk, full_name, file_type from chargemasters_files where year = {args.year} and file_ext = 'xlsx'"
        files = db_exec(conn, sql)

    dprint(f"files found # {len(files)}")

    for row in files:

        f = row['full_name']
        pk = row['pk']

        # f = "2011/El Camino Hosptial/106430763_Common25_2011.xlsx"
        # pk = 18

        print(f"file: {f}")

        try:

            wb = load_workbook(f, data_only=True)

            # if 'CDM All', then tab could be like 'HCAI 106381154_COMMON 25', '25 Most Common OP Procedures',
            #     'AB 1045 Form', 'AB 1045', 'AB 1045 FORM', or variation thereof.
            #
            if row['file_type'] != 'Common25':
                c25_sheet = common25_sheet(wb)
                if c25_sheet is not None:
                    wb.active = c25_sheet
                else:
                    continue
                    dprint(f"No Common25 sheet found in {f}")

            # It is ok to do these unconditionally, because it is safe when we start with no data.
            #
            db_exec(conn, f"update chargemasters_files set common25 = 0 where full_name = {fix(f)}")
            db_exec(conn, f"delete from chargemasters_common25 where file_pk = {row['pk']}")

            ws = wb.active

            found_top = False

            idx = 1

            while True:

                A = f"A{idx}"
                B = f"B{idx}"
                C = f"C{idx}"

                empty_rows = 0

                # print(f"checking row {idx}")

                if not found_top and is_cpt_label(ws[B]):
                    found_top = True

                if found_top and ws[A].value is None and ws[B].value is None and ws[C].value is None:
                    dprint("found empty line")
                    empty_lines += 1
                    if empty_lines > 10:
                        break
                else:
                    empty_lines = 0

                if found_top and ws[A].value is not None and ws[B].value is None and str(ws[C].value).startswith('='):
                    dprint("Formula found in |{ws[A].value}|{ws[B].value}|{ws[C].value}|")
                    break

                if not found_top and idx > 10:
                    dprint("No TOP cell found. Aborting.")
                    break

                if found_top and not str(ws[B].value).endswith('CPT Code'):
                    # print(f"proc: {ws[A].value}")
                    # print(f"code: {ws[B].value}")
                    # print(f"cost: {ws[C].value}")
                    # print("")


                    if ws[A].value is not None and ws[B].value is not None:
                        sql = f"""
                              insert into chargemasters_common25
                                  (file_pk, procedure_desc, cpt_code, charge_str)
                                  values
                                  ({pk}, {fix(ws[A].value)}, {fix(ws[B].value)}, {fix(ws[C].value)})
                              """
                        dprint(f"sql: {sql}")
                        try:
                            db_exec(conn, sql)
                        except:
                            print(f"A was |{ws[A].value}| and B was |{ws[B].value}|")
                            traceback.print_exc()

                idx += 1

            wb.close()

        except KeyboardInterrupt:
            print("quitting...")
            quit()
        except:
            traceback.print_exc()


def update_charge_value():

    sqls = [
        "update ignore chargemasters_common25 set charge = charge_str where charge_str rlike '^[0-9]*$'",
        "update ignore chargemasters_common25 set charge = charge_str where charge_str rlike '[0-9]*\\.[0-9]{1,2}'"
    ]

    for sql in sqls:
        try:
            db_exec(conn, sql)
        except KeyboardInterrupt:
            print("quitting...")
            quit()
        except:
            traceback.print_exc()


def update_file_counts():

    # TODO when I run this manually, it works but in this code, it returns everything. Why?
    # sql = "select charge_str as c from chargemasters_common25 where charge_str rlike '[0-9]*\\.[0-9]{3,}'"

    sql = "select charge_str as c from chargemasters_common25 where charge is NULL"
    for row in db_exec(conn, sql):
        if row['c'] is not None:
            # print(f"value: {row['c']}")
            parts = row['c'].split('.')
            if len(parts) == 2 and parts[1] == '':
                parts[1] = '00'
            if len(parts) == 2 and parts[0].isnumeric() and parts[1].isnumeric():
                next_charge = f"{parts[0]}.{parts[1][:2]}"
                sql = f"update chargemasters_common25 set charge = {next_charge} where charge_str = '{row['c']}'"
                dprint(f"sql: {sql}")
                try:
                    db_exec(conn, sql)
                except:
                    traceback.print_exc()

    sql = """
        update chargemasters_files
        set common25 = (select count(0) from chargemasters_common25 where file_pk = pk and charge is not NULL)
        where file_type = 'Common25';
    """
    try:
        db_exec(conn, sql)
    except KeyboardInterrupt:
        print("quitting...")
        quit()
    except:
        traceback.print_exc()

    print("")


if __name__ == '__main__':

    args = arguments()

    # if incremental, do not delete all data first.
    #
    incremental = args.file is not None or args.files is not None or args.year is not None

    if not args.update_charge and not args.update_files_counts:
        load_strings()

    if not args.only_load_strings and not args.update_files_counts:
        update_charge_value()

    if not args.only_load_strings and not args.update_charge:
        update_file_counts()

