from openpyxl import load_workbook
from xlrd import open_workbook


def load_my_workbook(ext, f):

    if ext == 'xlsx':
        return load_workbook(f)

    if ext == 'xls':
        return open_workbook(f)


def sheet_names(ext, wbook):
    if ext == 'xlsx':
        return wbook.sheetnames
    if ext == 'xls':
        raise Exception("what")


def only_sheet(ext, wbook):
    if ext == 'xlsx':
        return wbook.active
    if ext == 'xls':
        return wbook.sheet_by_index(0)
    return None


cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']


def find_common25_col_heads(ext, wbook, sheet_name):

    if ext != 'xlsx':
        return None

    found = None

    ws = wbook[sheet_name]

    for row in range(1,10):
        for col_idx in range(len(cols)-2):

            key1 = f"{cols[col_idx]}{(row)}"
            key2 = f"{cols[col_idx+1]}{(row)}"
            key3 = f"{cols[col_idx+2]}{(row)}"
            # print(f"keys: {key1}, {key2}, {key3}")

            col_names = {key1: ws[key1].value, key2: ws[key2].value, key3: ws[key3].value}
            if None not in col_names.values():
                return col_names
            # else:
            #     print(f"NOT {col_names}")

    return None


def cell(ext, ws, key):
    if ext == 'xlsx':
        return ws[key]
    return None


def cell_type(ext, wbook, ws, key):
    t = cell(ext, wbook[ws], key)
    if t is None:
        return None
    else:
        return t.data_type


def cell_value(ext, ws, key):
    if ext == 'xlsx':
        return ws[key].value
    return None


def cell_value_str(ext, wbook, sheet_name, key):

    if ext == 'xlsx':
        ws = wbook[sheet_name]
        if ws[key].value is None:
            return None
        else:
            return str(ws[key].value)

    if ext == 'xls':
        row = key[1:]
        col = key[0]
        if ws.cell_value(row, col) is None:
            return None
        else:
            return str(ws.cell_value(row, col))

    return None

