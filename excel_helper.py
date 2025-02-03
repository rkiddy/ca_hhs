import os
from datetime import datetime as dt
from pprint import pprint

from openpyxl import load_workbook
from sqlalchemy import create_engine

import config
import sql_helper
import traceback

def find_column_tops(wbook, sheet_num=None):

    tops = dict()

    for name in sheet_names(wbook):
        tops[name] = dict()

        for col in cols:
            for num in range(10):
                row = num + 1
                key = f"{col}{row}"
                val = cell_value_str(wbook, name, key)
                tops[name][f"{col}{row}"] = val
    return tops


def row_values_starting(ws, location, length=None):
    """Read values out of an xlsx row starting with the a sheet and a cell
       location (like 'A1' or 'D2') and going on until an empty value is
       reached (if every cell is expeced to be filled) or an expected length
       that will accept None values in cells."""

    cols = list()

    count = 0

    ccell = cell(ws, location)
    cvalue = cell_value(ws, location)
    clocation = location

    done = cvalue is None or cvalue == ''

    while not done:

        if length is None:
            done = cvalue is None or cvalue == ''
        else:
            done = len(cols) >= length

        if not done:
            cols.append(cell_value_str(ws, clocation))

            ccell = cell_to_right_of(ws, clocation)
            clocation = ccell.coordinate
            cvalue = cell_value(ws, clocation)

    return cols


def find_common25_col_heads(wbook, sheet_num):

    found = None

    ws = wbook[sheet_name]

    for row in range(1,10):
        for col_idx in range(len(cols)-2):

            key1 = f"{cols[col_idx]}{(row)}"
            key2 = f"{cols[col_idx+1]}{(row)}"
            key3 = f"{cols[col_idx+2]}{(row)}"
            # print(f"keys: {key1}, {key2}, {key3}")

            col_names = {key1: ws[key1].value, key2: ws[key2].value, key3: ws[key3].value}
            if None not in col_names.values():
                return col_names
            # else:
            #     print(f"NOT {col_names}")

    return None


def cell(ws, key):
    return ws[key]


def cell_type(ws, key):
    t = cell(ws, key)
    if t is None:
        return None
    else:
        return t.data_type


def cell_value(ws, key):
    return ws[key].value


def cell_value_str(ws, key):

    if ws[key].value is None:
        return None
    else:
        return str(ws[key].value)


def coordinate_parts(c):
    """Utility function for separating the two parts of an excel location,
       such as 'A5', or 'XYZ12'."""

    letters = list()
    numbers = list()
    for ltr in list(c):
        if ltr.isalpha():
            letters.append(ltr)
        else:
            numbers.append(ltr)
    ltrs = ''.join(letters)
    num = int(''.join(numbers))
    return [ltrs, num]


def column_offset(col, inc):
    """Given a column part of a excel cell location, such as 'A' or 'AF', return the one
       to the right (inc = 1) or the left (inc = -1)."""
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    parts = list(col)
    if inc == 1:
        # handle the easy case
        if parts[-1] in list(alpha)[:-1]:
            val = ord(parts[-1])
            nval = chr(val+1)
            parts[-1] = nval
            return ''.join(parts)
        # adding from Z.
        if parts[-1] == 'Z':
            parts[-1] = 'A'
            parts.append('A')
            return ''.join(parts)
    if inc == -1:
        # handle the easy case
        if parts[-1] in list(alpha)[1:]:
            val = ord(parts[-1])
            nval = chr(val-1)
            parts[-1] = nval
            return ''.join(parts)
        if parts[-1] == 'A' and len(parts) > 1:
            if parts[-2] == 'A':
                parts.pop()
                parts[-1] = 'Z'
            else:
                parts.pop()
                val = ord(parts[-1])
                nval = chr(val-1)
                parts[-1] = nval
                parts.append('Z')
            return ''.join(parts)

    raise Exception(f"invalid column name passed: {col}, inc: {col}")


def cell_to_right_of(ws, key):
    parts = coordinate_parts(key)
    ncol = column_offset(parts[0], 1)
    return ws[f"{ncol}{parts[1]}"]


def cell_to_left_of(ws, key):
    parts = coordinate_parts(key)
    ncol = column_offset(parts[0], -1)
    return ws[f"{ncol}{parts[1]}"]


def cell_above(ws, key):
    parts = coordinate_parts(key)
    nrow = parts[1] - 1
    return ws[f"{parts[0]}{nrow}"]


def cell_below(ws, key):
    parts = coordinate_parts(key)
    nrow = parts[1] + 1
    return ws[f"{parts[0]}{nrow}"]


def all_empty(values):
    not_empty = [r for r in values if r is not None and r != '']
    return len(not_empty) == 0


def location_of(c):
    return c.coordinate


# TODO the create_tables and read_tables can be refactored to find once and produce different statements.
#
def create_tables(tables, types={}, replaces={}, length_pad=0, verbose=False, topmost_data=2, leftmost_data='A'):
    """Do the right thing for most csv files based on the column header strings found.
       Also determines the optimal length for a string column, though this process reads
       all of the data and so is probably not optimal and is probably doable in a smarter
       way."""

    if verbose:
        print("create_tables:: starting")

    for table in tables:

        file_name = tables[table]['file']
        sheet_name = tables[table]['sheet']

        print(f"file: {file_name}, sheet: {sheet_name}")
        try:

            wb = load_workbook(file_name)

            if sheet_name:
                for sheet in wb.sheetnames:
                    if sheet == sheet_name:
                        wb.active = wb.get_sheet_by_name(sheet)

            ws = wb.active

            first_col_head = f"{leftmost_data}1"

            # Find the column tops.
            #
            # TODO: Finding column tops may not be as easy as going to location A1.
            #
            # Turn the list of columns into a dictionary with column names going to 0.
            #
            cols = sql_helper.fix_col_heads(row_values_starting(ws, first_col_head))
            cols = dict(zip(cols, [0] * len(cols)))

            if verbose:
                print(f"cols: {cols}")

            done = False

            idx = topmost_data

            # Read through content, row at a time, noting max lengths of string values.
            #
            while not done:

                current = f"{leftmost_data}{idx}"
                values = row_values_starting(ws, current, length=len(cols))

                if verbose:
                    print(f"values: {values}")

                if all_empty(values):

                    done = True
                    if verbose:
                        print(f"all_empty() returns True at idx: {idx}")

                else: # not yet finished.

                    for jdx in range(len(cols)):
                        if values[jdx] is not None and len(values[jdx]) > cols[list(cols.keys())[jdx]]:
                            cols[list(cols.keys())[jdx]] = len(values[jdx])

                    idx += 1

            if verbose:
                print(f"cols: {cols}")

            col_defs = [f"{col_name} varchar({cols[col_name]})" for col_name in cols]

            sql = f"create table {table} ("
            sql += ', '.join(col_defs)
            sql += ")"

            print(f"\ncreating table {table}...")

            sql_helper.db_exec_sql(f"drop table if exists {table}")

            sql_helper.db_exec_sql(sql)

            wb.close()

        except:
            traceback.print_exc()


def read_data(tables, types={}, replaces={}, start_row=None, bucket=1000, verbose=False, topmost_data=2, leftmost_data='A'):
    """Reads CSV data and puts it into a table."""

    if verbose:
        print("read_data:: starting")

    # print(f"read_data: types: {types}")

    for table in tables:

        try:

            file_name = tables[table]['file']
            sheet_name = tables[table]['sheet']

            print(f"\nfile: {file_name}, sheet: {sheet_name} -> ", end='')

            wb = load_workbook(file_name)

            if sheet_name:
                for sheet in wb.sheetnames:
                    if sheet == sheet_name:
                        wb.active = wb.get_sheet_by_name(sheet)
            ws = wb.active

            first_col_head = f"{leftmost_data}1"

            # Find the column tops.
            #
            cols = sql_helper.fix_col_heads(row_values_starting(ws, first_col_head))
            cols = dict(zip(cols, [0] * len(cols)))

            prefix = f"insert into {table} ({', '.join(cols)}) values"

            suffixes = list()

            if verbose:
                print(f"prefix: {prefix}")

            done = False

            idx = topmost_data

            added = 0

            # Read through content, row at a time.
            #
            while not done:

                current = f"{leftmost_data}{idx}"
                values = row_values_starting(ws, current, length=len(cols))

                if verbose:
                    print(f"values: {values}")

                if all_empty(values):

                    done = True
                    if verbose:
                        print(f"all_empty() returns True at idx: {idx}")

                else: # not yet finished.

                    values = [sql_helper.fix(r) for r in values]

                    suffix = f"({', '.join(values)})"

                    suffixes.append(suffix)

                    if len(suffixes) > bucket:
                        sql_helper.db_exec_many_sql(prefix, suffixes)
                        added += len(suffixes)
                        suffixes = list()

                    idx += 1

            if len(suffixes) > 0:
                added += len(suffixes)
                sql_helper.db_exec_many_sql(prefix, suffixes)

            print(f"table: {table} # {added}")

            wb.close()
        except:
            traceback.print_exc()


if __name__ == '__main__':
    pass

