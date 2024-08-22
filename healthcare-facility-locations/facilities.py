import csv
import os
import sys

from openpyxl import load_workbook
from sqlalchemy import create_engine

sys.path.append('..')

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def db_exec(eng, this_sql):
    # print(f"sql: {sql}")
    if this_sql.strip().startswith('select'):
        return [dict(row) for row in eng.execute(this_sql).fetchall()]
    else:
        # print(f"sql: {this_sql}")
        return eng.execute(this_sql)


def fix_num(start):
    return start.replace(',', '')


def fix_value(start):
    return str(start).replace("'", "''").replace('%', '%%')


if __name__ == '__main__':

    db_exec(conn, "drop table if exists healthcare_facilities")

    sql = f"""
    create table healthcare_facilities (
         licensed_certified varchar(255),
         flag varchar(255),
         t18_19 varchar(31),
         facid char(9),
         fac_status_type_code varchar(31),
         aspen_facid varchar(255),
         ccn varchar(31), # integer,
         terminat_sw char(1),
         participation_date varchar(255),
         approval_date varchar(255),
         npi varchar(11), # integer,
         can_be_deemed_fac_type char(1),
         can_be_certified_fac_type char(1),
         deemed varchar(255),
         ao_cd varchar(255),
         dmg_efctv_dt varchar(255),
         ao_trmntn_dt varchar(255),
         ao_name varchar(255),
         facname varchar(255),
         fac_type_code varchar(255),
         fac_fdr varchar(255),
         ltc varchar(255),
         capacity varchar(11), # integer,
         address varchar(255),
         city varchar(255),
         zip varchar(9),
         zip9 varchar(9),
         facadmin varchar(255),
         contact_email varchar(255),
         contact_fax varchar(15), # bigint,
         contact_phone_number varchar(15), # bigint,
         county_code varchar(255),
         county_name varchar(255),
         district_number varchar(11), # integer,
         district_name varchar(255),
         isfacmain char(1),
         parent_facid varchar(9),
         fac_fac_relationship_type_code varchar(255),
         start_date varchar(255),
         license_number varchar(255),
         business_name varchar(255),
         license_status_description varchar(255),
         initial_license_date varchar(255),
         license_effective_date varchar(255),
         license_expiration_date varchar(255),
         entity_type_description varchar(255),
         latitude decimal(12,8),
         longitude decima(12,8),
         location varchar(255),
         hcai_id varchar(9),
         cclho_code varchar(255),
         cclho_name varchar(255),
         fips_county_code varchar(255),
         birthing_facility_flag varchar(255),
         trauma_ped_ctr varchar(255),
         trauma_ctr varchar(255),
         type_of_care varchar(255),
         critical_access_hospital varchar(255),
         data_date varchar(255));"""

    nums = ['ccn', 'npi', 'capacity', 'district_number']

    db_exec(conn, sql)

    f = "licensed-and-certified-healthcare-facility-locations.xlsx"

    wb = load_workbook(f)

    sheet_ok = False

    row_num = 0

    for row in wb.active.iter_rows():

        row_num += 1

        if row_num == 1:
            continue

        data = list()

        # if row[0].row > 10:
        #     quit()

        # print(f"row: {row}")

        values = [f"'{fix_value(r.value)}'" for r in row]

        sql = f"insert into healthcare_facilities values ({', '.join(values)})"
        # print(f"sql: {sql}")
        db_exec(conn, sql)

    # db_exec(conn, "alter table healthcare_facilities change column ccn ccn bigint")
    # db_exec(conn, "alter table healthcare_facilities change column npi npi bigint")

    db_exec(conn, "update healthcare_facilities set capacity = NULL where capacity = ''")
    db_exec(conn, "alter table healthcare_facilities change column capacity capacity bigint")

    db_exec(conn, "update healthcare_facilities set district_number = NULL where district_number = ''")
    db_exec(conn, "alter table healthcare_facilities change column district_number district_number bigint")

    db_exec(conn, "drop table if exists healthcare_facilities_across_time")

    sql = """
          create table healthcare_facilities_across_time (
              state_fiscal_year char(10),
              county_name varchar(63),
              provider_type varchar(127),
              fac_fdr varchar(127),
              count integer)"""

    db_exec(conn, sql)

    f = "across-time-summary-data-licensed-and-certified-healthcare-facilities.csv"

    with open(f, newline='') as csvfile:
        rdr = csv.DictReader(csvfile)

        for row in rdr:
            # print(f"row: {row}")

            values = list(row.values())

            values[0] = f"'{fix_value(values[0])}'"
            values[1] = f"'{fix_value(values[1])}'"
            values[2] = f"'{fix_value(values[2])}'"
            values[3] = f"'{fix_value(values[3])}'"

            sql = f"insert into healthcare_facilities_across_time values ({', '.join(values)})"
            # print(f"sql: {sql}")
            db_exec(conn, sql)

    sql = "insert into updates values ('healthcare-facility-locations', unix_timestamp(now()))"
    db_exec(conn, sql)

