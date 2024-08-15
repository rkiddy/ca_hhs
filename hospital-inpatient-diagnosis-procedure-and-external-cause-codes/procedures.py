
import os
import traceback

from openpyxl import load_workbook
from sqlalchemy import create_engine

sys.path.append('..')

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def db_exec(engine, sql):
    # print(f"sql: {sql}")
    if sql.strip().startswith('select'):
        return [dict(r) for r in engine.execute(sql).fetchall()]
    else:
        return engine.execute(sql)


def create_tables():
    tmp_sqls = list()

    tmp_sqls.append("drop table if exists inpatient_deaths;")
    tmp_sqls.append("""
       create table inpatient_deaths (
         icdcm_code varchar(31),
         diagnosis_description varchar(1027),
         total int,
         year int)""")

    tmp_sqls.append("drop table if exists inpatient_death_causes;")
    tmp_sqls.append("""
       create table inpatient_death_causes (
         icdcm_code varchar(31) primary key,
         diagnosis_description varchar(1027))""")

    tmp_sqls.append("drop table if exists inpatient_diagnoses;")
    tmp_sqls.append("""
       create table inpatient_diagnoses (
         icdcm_code varchar(31),
         diagnosis_description varchar(1027),
         total int,
         year int)""")

    tmp_sqls.append("drop table if exists inpatient_procedures;")
    tmp_sqls.append("""
       create table inpatient_procedures (
         icdpcs_code varchar(31),
         procedure_description varchar(1027),
         total int,
         year int)""")

    return tmp_sqls


def get_sheet_name(file_part):
    if file_part == 'diagnosis':
        return 'ICD-10-CM'
    if file_part == 'morbidity':
        return 'ICD-10-CM'
    if file_part == 'procedure':
        return 'ICD-10-PCS'


def get_table_name(file_part):
    if file_part == 'diagnosis':
        return 'inpatient_diagnosis'
    if file_part == 'morbidity':
        return 'inpatient_deaths'
    if file_part == 'procedure':
        return 'inpatient_procedures'


def get_columns(file_part):
    if file_part == 'diagnosis':
        cols = ['icdcm_code', 'diagnosis_description']
    if file_part == 'procedure':
        cols = ['icdpcs_code', 'procedure_description']
    if file_part == 'morbidity':
        cols = ['icdcm_code', 'diagnosis_description']
    cols.extend(['total', 'year'])
    return cols


def is_int_column(cname):
    return cname == 'total' or cname == 'year'


def fix_value(val):
    return val.replace("'", "''").replace('%', '%%').replace('_x000D_', '').replace('\n', ' ').upper()


if __name__ == '__main__':

    for sql in create_tables():
        db_exec(conn, sql)

    for f in os.listdir('.'):

        sqls = list()

        if not str(f).endswith('.xlsx'):
            continue

        parts = str(f).split('-')

        year = int(parts[0])

        if year < 2017:
            continue

        # if parts[-3] != 'morbidity':
        #     continue

        columns = get_columns(parts[-3])
        table_name = get_table_name(parts[-3])

        print(f"file: {f}, year: {year}, table: {table_name}")

        wb = load_workbook(f)

        sheet_ok = False

        # set the correct sheet to active
        for found_sheet in wb.sheetnames:
            if found_sheet.startswith('ICD'):
                wb.active = wb[found_sheet]
                sheet_ok = True

        if not sheet_ok:
            print(f"Could not find proper sheet in list: {wb.sheetnames}")
            continue

        first_row = True
        row_num = 0

        for row in wb.active.iter_rows():

            row_num += 1

            data = list()

            if first_row:
                first_row = False
                continue

            data.append(str(row[0].value))
            data.append(str(row[1].value))
            data.append(str(row[2].value))

            if data[-1] is None or data[-1] == 'None':
                data[-1] = 'NULL'
            else:
                if not data[-1].isnumeric():
                    print(f"row # {row_num} has non-numeric total: {data[-1]}")
                    continue

            data.append(year)

            cdata = dict(zip(columns, data))

            for col in columns:

                if not is_int_column(col):
                    value = str(fix_value(cdata[col]))
                    cdata[col] = f"'{value}'"
                else:
                    cdata[col] = str(cdata[col])
                    if cdata[col] == 'None' or cdata[col] == "'None'":
                        cdata[col] = 'NULL'

            data = list(cdata.values())

            sqls.append(f"insert into {table_name} values ({', '.join(data)})")

        if len(sqls) > 0:

            print(f"rows #: {len(sqls)}")

            num = 0

            for sql in sqls:
                # print(f"{sql};")
                try:
                    db_exec(conn, sql)
                    num += 1
                    if (num % 1000) == 0:
                        print(f"{num}...")
                except:
                    print(f"row num # {row_num}")
                    traceback.print_exc()

    sql = "insert ignore into inpatient_death_causes " \
          "(select icdcm_code, diagnosis_description from inpatient_deaths order by icdcm_code, diagnosis_description)"
    db_exec(conn, sql)

    sql = "alter table inpatient_deaths drop column diagnosis_description"
    db_exec(conn, sql)

    sql = """
          insert into updates values ('hospital-inpatient-diagnosis-procedure-and-external-cause-codes',
          unix_timestamp(now()))"""

    db_exec(conn, sql)

